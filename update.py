import time
import os
import openpyxl
from pathlib import Path
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

EXCEL_FILE_NAME = "varlik.xlsx"
SHEET_NAME = "sheet"

def get_data():
    url_market = "https://www.altinkaynak.com/canli-kurlar/"
    url_bank = "https://www.yapikredi.com.tr/yatirimci-kosesi/altin-bilgileri"
    scraped_data = []

    chrome_options = Options()
    chrome_options.add_argument("--headless") 
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    
    print("Starting browser...")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    
    try:
        print(f"Loading {url_market}...")
        driver.get(url_market)
        time.sleep(5) 

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")

        def extract_from_table(keyword):
            target_table = None
            for table in soup.find_all("table", class_="table"):
                headers = [th.get_text(strip=True) for th in table.find_all("th")]
                if keyword in headers:
                    target_table = table
                    break
            
            if target_table:
                tbody = target_table.find("tbody")
                if tbody:
                    rows = tbody.find_all("tr")
                    for row in rows:
                        cols = row.find_all("td")
                        if len(cols) >= 3:
                            name = cols[0].get_text(strip=True)
                            buy = cols[1].get_text(strip=True)
                            scraped_data.append((name, buy))

        extract_from_table("Döviz")
        extract_from_table("Altın")

        print(f"Loading {url_bank}...")
        driver.get(url_bank)
        time.sleep(5) 
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")
        value = soup.select_one("#leftBoxResult span").text.strip().replace("  TL", "").strip()

        scraped_data.append(("Banka Altın", value))

        return scraped_data
    except Exception as e:
        print(f"Scraping error: {e}")
        return []
    finally:
        driver.quit()

def save_to_excel(data_list, filename):
    
    if not os.path.exists(filename):
        print(f"The file '{filename}' does not exist. Creating file.")
        if not Path(EXCEL_FILE_NAME).exists():
            openpyxl.Workbook().save(Path(EXCEL_FILE_NAME))

    try:
        print(f"Opening {filename}...")
        workbook = openpyxl.load_workbook(filename)
        
        if SHEET_NAME in workbook.sheetnames:
            sheet = workbook[SHEET_NAME]
        else:
            sheet = workbook.active
            
        sheet["A1"] = "Name"
        sheet["B1"] = "Buy Price"

        for i, (name, buy_val) in enumerate(data_list, start=2):
            
            sheet[f"A{i}"] = name
            
            try:
                clean_val = float(buy_val.replace('.', '').replace(',', '.'))
                sheet[f"B{i}"] = clean_val
            except ValueError:
                sheet[f"B{i}"] = buy_val

        workbook.save(filename)
        print(f"Successfully updated")

    except Exception as e:
        print(f"Excel Error: {e}")

if __name__ == "__main__":
    data = get_data()
    if data:
        save_to_excel(data, EXCEL_FILE_NAME)
    else:
        print("No data found to save.")